from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.http.response import HttpResponse
from django.db.models import Q ## ESTO SIRVE PARA HACER BÚSQUEDA
from datetime import date, timedelta
import json
from openpyxl import Workbook
from .models import Pedido, Variante, Articulo, Estado, Servicio, pedido_variante
#from apps.seguridad.models import Usuario 
from scripts.utilidades_db import guarda_citas, guarda_articulos
# Create your views here.
@login_required
def home(request): #acercade
    if request.method == 'POST':
        if request.POST['nombre'] == 'sinc':
            FechaActual=request.POST['FechaInicial']
            FechaFinal=request.POST['FechaFinal']
            #print((FechaActual,FechaFinal))
            #print('----------')
            guarda_articulos()
            guarda_citas(FechaActual+'T00:00:00Z', FechaFinal+'T23:59:00Z')
            return redirect(f'./?initial={FechaActual}&final={FechaFinal}')
        elif request.POST['nombre'] == 'search':
            search=request.POST['search']
            FechaActual=request.POST['FechaInicial']
            FechaFinal=request.POST['FechaFinal']
            return redirect(f'./?initial={FechaActual}&final={FechaFinal}&search={search}')
        elif request.POST['nombre'] == 'addproduct':
            form=request.POST
            print(form)
            print('----------------') 
            valor_servicio=form.get('Servicio','1650382159084')
            if valor_servicio == 'on':
                valor_servicio='1658428294592'
            Confirmed_by_customer=form.get('Confirmed_by_customer')=='on'
            In_Production=form.get('In_production')=='on'
            datos_guardados=Pedido.objects.filter(IdPedidoSquare=form['IdPedido'])
            Cancel=form.get('Cancel') == 'on'
            datos_guardados.update(
                    Observacion=form['Observacion'],
                    Estado_id=form['Estado'],
                    Servicio_id=valor_servicio,
                    Cancelado=Cancel,
                    Confirmed_by_customer=Confirmed_by_customer,
                    In_Production=In_Production,
                    Direccion=form.get('pedido-direccion'))
            ## SE AÑADEN LOS PRODUCTOS
            prod_json=form.getlist('prod')
            para_borrar=pedido_variante.objects.filter(pedido_id=form['IdPedido']) 
            para_borrar.delete()
            for producto in prod_json:
                producto=json.loads(producto)
                pedido_variante.objects.create(
                    pedido_id=form['IdPedido'],
                    variante_id=producto['name'],
                    cantidad=producto['quantity'])
    Finicial=request.GET.get('initial')
    Ffinal=request.GET.get('final')
    search=request.GET.get('search')
    if Finicial:
        FechaActual=date.fromisoformat(Finicial)
    else:
        FechaActual = date.today()
    if Ffinal:
        FechaFinal=date.fromisoformat(Ffinal)
    else:
        FechaFinal = FechaActual + timedelta(1)
    servicios=Servicio.objects.all()
    estados=Estado.objects.filter(active=True)

    if search:
        pedidos=Pedido.objects.filter(
        Q(Fecha__gte = FechaActual)&
        Q(Fecha__lte = FechaFinal)&
        Q(NombreCliente__icontains=search) |
        Q(Telefono__icontains=search)
        ).distinct().order_by('Fecha','Hora','IdPedidoSquare')
        ctx={
            'pedidos':pedidos,
            'search':search,
            'servicios':servicios,
            'estados':estados}
    else:
        if request.user.admin_app:
            pedidos=Pedido.objects.filter(
            Fecha__gte=FechaActual,
            Fecha__lte=FechaFinal).order_by('Fecha','Hora','IdPedidoSquare')
        else:
            pedidos=Pedido.objects.filter(
                Servicio_id__Tipo='D',
                Cancelado=False,
                Fecha__gte=FechaActual,
                Fecha__lte=FechaFinal).order_by('Fecha','Hora','IdPedidoSquare') 
        search=''
        ctx={
            'pedidos':pedidos,
            'search':search,
            'servicios':servicios,
            'estados':estados}
    return render(request,'venta/home.html',ctx)

@login_required
def api_productos(request):
    productos=Variante.objects.all()
    lista_productos=[]
    for producto in productos:
        data_producto={}
        data_producto['value']=producto.IdVarianteSquare
        Name=Articulo.objects.get(IdArticuloSquare = producto.IdArticulo_id)
        data_producto['label']=Name.Descripcion+' - '+producto.Descripcion
        lista_productos.append(data_producto)
    IdPedido=request.GET.get('idpedido')
    ctx={'products':lista_productos}
    if IdPedido:
        ContenidoPedido=Pedido.objects.get(IdPedidoSquare=IdPedido)
        pedidos=list(pedido_variante.objects.filter(pedido_id=IdPedido).values('cantidad','variante_id__IdArticulo__Descripcion','variante_id__Descripcion','variante_id'))
        #print(pedidos)
        ValorServicio=str(ContenidoPedido.Servicio) =='D'
        #print(type(ContenidoPedido.Servicio))
        #print(ValorServicio)
        #print('----')
        ctx={
                'Observacion':ContenidoPedido.Observacion,
                'Notas':ContenidoPedido.Notas,
                'NombreCliente':ContenidoPedido.NombreCliente,
                'Telefono':ContenidoPedido.Telefono,
                'Direccion':ContenidoPedido.Direccion,
                'Confirmed_by_customer':ContenidoPedido.Confirmed_by_customer,
                'In_Production':ContenidoPedido.In_Production,
                'FechaHora':ContenidoPedido.Fecha.strftime("%d/%m/%y")+' - ' +ContenidoPedido.Hora.strftime("%H:%m"),
                'Cancelado':ContenidoPedido.Cancelado,
                'Estado':ContenidoPedido.Estado_id,
                'Servicio':ValorServicio,
                'ListaPedidos':pedidos
            }
    return JsonResponse(ctx)


@login_required
def ExportaExcel(request):
    if request.method == 'POST':
        form=request.POST
        ids=list(form.getlist('pedidos'))
        lista_pedidos=[]
        for IdPedido in ids:
            ContenidoPedido=Pedido.objects.get(IdPedidoSquare=IdPedido)
            pedidos=list(pedido_variante.objects.filter(pedido_id=IdPedido).values('cantidad','variante_id__IdArticulo__Descripcion','variante_id__Descripcion','variante_id'))
            if ContenidoPedido.Servicio_id == '1658428294592':
                servicio='Delivery'
            else:
                servicio='No Delivery'
            pedido={
                'IdPedido':IdPedido,
                'Observacion':ContenidoPedido.Observacion,
                'Notas':ContenidoPedido.Notas,
                'NombreCliente':ContenidoPedido.NombreCliente,
                'Telefono':ContenidoPedido.Telefono,
                'Direccion':ContenidoPedido.Direccion,
                'Fecha':ContenidoPedido.Fecha.strftime("%d/%m/%y"),
                'Hora':ContenidoPedido.Hora.strftime("%H:%m"),
                'Cancelado':ContenidoPedido.Cancelado,
                'Estado':ContenidoPedido.Estado_id,
                'Servicio':servicio,
                'ListaPedidos':pedidos
            }
            lista_pedidos.append(pedido)
            
        cont=2
        wb=Workbook()
        ws=wb.active
        ws['A1']='Date'
        ws['B1']='Status'
        ws['C1']='Time'
        ws['D1']='Service'
        ws['E1']='Cancel'
        ws['F1']='Client'
        ws['G1']='Phone'
        ws['H1']='Address'
        ws['I1']='Notes'
        ws['J1']='Product'
        ws['K1']='Quantity'
        ws['L1']='Obs'
        for pedido in lista_pedidos: 
            variantes=pedido.get('ListaPedidos')
            if len(variantes)==0:
                variantes=[{'variante_id__Descripcion':'','cantidad':'','variante_id__IdArticulo__Descripcion':'' }]
            for variante in variantes:
                print(variante)
                print('-----------')
                ws.cell(row=cont,column=1).value=pedido['Fecha']
                ws.cell(row=cont,column=2).value=pedido['Estado']
                ws.cell(row=cont,column=3).value=pedido['Hora']
                ws.cell(row=cont,column=4).value=pedido['Servicio']
                ws.cell(row=cont,column=5).value=pedido['Cancelado']
                ws.cell(row=cont,column=6).value=pedido['NombreCliente']
                ws.cell(row=cont,column=7).value=pedido['Telefono']
                ws.cell(row=cont,column=8).value=pedido['Direccion']
                ws.cell(row=cont,column=9).value=pedido['Notas']
                ws.cell(row=cont,column=10).value=variante['variante_id__IdArticulo__Descripcion']+' - '+variante['variante_id__Descripcion']
                ws.cell(row=cont,column=11).value=variante['cantidad']
                ws.cell(row=cont,column=12).value=pedido['Observacion']
                cont+=1
        nombre_archivo="ReportePedidos.xlsx"
        response= HttpResponse(content_type='application/ms-excel')
        content="attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
    return response
